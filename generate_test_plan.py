#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tạo file Excel tiến độ test các chức năng của từng microservice
trong dự án e-commerce-microservices
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
import re

# ======================================================
# DỮ LIỆU CHỨC NĂNG TỪNG SERVICE (STT, Chức năng, API Endpoint, Mô tả)
# ======================================================
SERVICES_DATA = [
    {
        "service": "user-service",
        "color": "4472C4",  # xanh dương
        "features": [
            # AuthController
            ("1", "Đăng nhập bằng username/password", "POST /auth/login", "Xác thực người dùng, trả JWT token. Hỗ trợ 2FA & OTP"),
            ("2", "Đăng nhập không mật khẩu (Passwordless)", "POST /auth/passwordless/request", "Gửi OTP qua email để đăng nhập không cần mật khẩu"),
            ("3", "Xác minh thiết bị đăng nhập", "GET /auth/device-approval/verify", "Xác minh token thiết bị mới từ email"),
            ("4", "Xác minh OTP đăng nhập", "POST /auth/login-otp/verify", "Xác minh mã OTP gửi qua email khi đăng nhập"),
            ("5", "Xác minh 2FA (TOTP)", "POST /auth/login-2fa/verify", "Xác minh mã TOTP từ ứng dụng xác thực"),
            ("6", "Gửi lại OTP đăng nhập", "POST /auth/login-otp/resend", "Gửi lại OTP đăng nhập khi OTP hết hạn"),
            ("7", "Làm mới Access Token", "POST /auth/refresh", "Dùng Refresh Token để lấy Access Token mới"),
            ("8", "Quên mật khẩu", "POST /auth/password/forgot", "Gửi email đặt lại mật khẩu"),
            ("9", "Đặt lại mật khẩu", "POST /auth/password/reset", "Đặt lại mật khẩu bằng token từ email"),
            # RegisterController
            ("10", "Đăng ký tài khoản mới", "POST /registration", "Tạo tài khoản, gửi OTP xác thực qua email"),
            ("11", "Xác minh OTP đăng ký", "GET /registration/verify", "Kích hoạt tài khoản bằng OTP hoặc token"),
            ("12", "Kiểm tra username/email trùng", "GET /registration/check", "Kiểm tra username/email đã tồn tại chưa"),
            ("13", "Gửi lại OTP kích hoạt", "POST /registration/resend-otp", "Gửi lại OTP kích hoạt tài khoản"),
            # UserController
            ("14", "Lấy danh sách người dùng", "GET /users", "Lấy tất cả người dùng (Admin)"),
            ("15", "Tìm người dùng theo tên", "GET /users?name=", "Tìm kiếm người dùng theo username"),
            ("16", "Lấy người dùng theo ID", "GET /users/{id}", "Lấy thông tin chi tiết người dùng theo ID"),
            ("17", "Lấy thiết bị đăng nhập", "GET /users/{id}/devices", "Danh sách thiết bị đăng nhập của người dùng"),
            ("18", "Lịch sử thay đổi hồ sơ", "GET /users/{id}/profile-changes", "Xem log thay đổi thông tin hồ sơ"),
            ("19", "Tạo người dùng (Admin)", "POST /users", "Tạo người dùng mới (Admin tạo thay)"),
            ("20", "Cập nhật vai trò người dùng", "PUT /users/{id}/role", "Thay đổi role của người dùng"),
            ("21", "Mở khóa tài khoản", "PUT /users/{id}/unlock", "Admin mở khóa tài khoản bị khóa"),
            ("22", "Lấy danh sách địa chỉ", "GET /users/{id}/addresses", "Danh sách địa chỉ giao hàng của người dùng"),
            ("23", "Lấy địa chỉ mặc định", "GET /users/{id}/address", "Lấy địa chỉ giao hàng mặc định"),
            ("24", "Thêm địa chỉ giao hàng", "POST /users/{id}/addresses", "Thêm địa chỉ giao hàng mới"),
            ("25", "Cập nhật địa chỉ giao hàng", "PUT /users/{id}/addresses/{addressId}", "Cập nhật địa chỉ giao hàng"),
            ("26", "Xóa địa chỉ giao hàng", "DELETE /users/{id}/addresses/{addressId}", "Xóa địa chỉ giao hàng"),
            ("27", "Đặt địa chỉ mặc định", "PATCH /users/{id}/addresses/{addressId}/default", "Đặt địa chỉ làm mặc định"),
            ("28", "Lấy danh sách tỉnh/thành VN", "GET /addresses/vn/provinces", "Lấy danh sách tỉnh thành Việt Nam"),
            ("29", "Lấy danh sách phường/xã", "GET /addresses/vn/provinces/{code}/wards", "Lấy danh sách phường/xã theo tỉnh"),
            # UserProfileController
            ("30", "Xem hồ sơ người dùng", "GET /users/{id}/profile", "Lấy thông tin hồ sơ cá nhân"),
            ("31", "Cập nhật hồ sơ cá nhân", "PUT /users/{id}/profile", "Cập nhật thông tin hồ sơ"),
            ("32", "Kiểm tra email tồn tại", "GET /users/check-email", "Kiểm tra email đã được dùng chưa"),
            ("33", "Gửi OTP thay đổi email", "POST /users/{id}/email/otp", "Gửi OTP để xác minh email mới"),
            ("34", "Xác nhận thay đổi email", "POST /users/{id}/email/confirm", "Xác nhận thay đổi email bằng OTP"),
            ("35", "Gửi lại OTP thay đổi email", "POST /users/{id}/email/resend-otp", "Gửi lại OTP đổi email"),
            ("36", "Đổi mật khẩu", "PATCH /users/{id}/password", "Thay đổi mật khẩu người dùng"),
            ("37", "Xác minh mật khẩu hiện tại", "POST /users/{id}/password/verify", "Xác minh mật khẩu hiện tại trước khi đổi"),
            ("38", "Tạo mã QR 2FA", "POST /users/{id}/2fa/generate", "Tạo secret và QR code cho 2FA"),
            ("39", "Bật 2FA", "POST /users/{id}/2fa/enable", "Bật xác thực 2 bước (TOTP)"),
            ("40", "Tắt 2FA", "POST /users/{id}/2fa/disable", "Tắt xác thực 2 bước"),
            ("41", "Upload ảnh đại diện", "POST /users/{id}/avatar", "Upload avatar người dùng"),
            ("42", "Lấy file ảnh đại diện", "GET /users/{id}/avatar/file/{fileName}", "Lấy file ảnh đại diện"),
        ]
    },
    {
        "service": "product-catalog-service",
        "color": "70AD47",  # xanh lá
        "features": [
            ("1", "Lấy danh sách sản phẩm (Public)", "GET /products", "Lấy danh sách sản phẩm công khai, hỗ trợ filter & phân trang"),
            ("2", "Lấy chi tiết sản phẩm (Public)", "GET /products/{id}", "Lấy thông tin chi tiết sản phẩm"),
            ("3", "Lấy biến thể sản phẩm (Public)", "GET /products/{id}/variants", "Lấy danh sách biến thể của sản phẩm"),
            ("4", "Lấy chi tiết biến thể (Public)", "GET /variants/{id}", "Lấy thông tin biến thể theo ID"),
            ("5", "Lấy danh sách danh mục (Public)", "GET /categories", "Lấy cây danh mục sản phẩm"),
            ("6", "Lấy chi tiết danh mục (Public)", "GET /categories/{id}", "Lấy thông tin danh mục theo ID"),
            ("7", "Lấy danh sách thương hiệu (Public)", "GET /brands", "Lấy danh sách thương hiệu"),
            ("8", "Lấy chi tiết thương hiệu (Public)", "GET /brands/{id}", "Lấy thông tin thương hiệu"),
            ("9", "Lấy log thay đổi sản phẩm", "GET /products/{id}/changelogs", "Lịch sử thay đổi sản phẩm"),
            # Admin Product
            ("10", "Lấy sản phẩm (Admin)", "GET /admin/products", "Lấy danh sách sản phẩm trang Admin"),
            ("11", "Tạo sản phẩm (Admin)", "POST /admin/products", "Tạo sản phẩm mới"),
            ("12", "Cập nhật sản phẩm (Admin)", "PUT /admin/products/{id}", "Cập nhật thông tin sản phẩm"),
            ("13", "Xóa sản phẩm (Admin)", "DELETE /admin/products/{id}", "Xóa sản phẩm"),
            ("14", "Cập nhật trạng thái sản phẩm", "PATCH /admin/products/{id}/status", "Bật/tắt hiển thị sản phẩm"),
            # Admin Variant
            ("15", "Tạo biến thể sản phẩm (Admin)", "POST /admin/products/{id}/variants", "Tạo biến thể mới cho sản phẩm"),
            ("16", "Cập nhật biến thể (Admin)", "PUT /admin/variants/{id}", "Cập nhật thông tin biến thể"),
            ("17", "Xóa biến thể (Admin)", "DELETE /admin/variants/{id}", "Xóa biến thể sản phẩm"),
            # Admin Category
            ("18", "Tạo danh mục (Admin)", "POST /admin/categories", "Tạo danh mục mới"),
            ("19", "Cập nhật danh mục (Admin)", "PUT /admin/categories/{id}", "Cập nhật danh mục"),
            ("20", "Xóa danh mục (Admin)", "DELETE /admin/categories/{id}", "Xóa danh mục"),
            # Admin Brand
            ("21", "Tạo thương hiệu (Admin)", "POST /admin/brands", "Tạo thương hiệu mới"),
            ("22", "Cập nhật thương hiệu (Admin)", "PUT /admin/brands/{id}", "Cập nhật thông tin thương hiệu"),
            ("23", "Xóa thương hiệu (Admin)", "DELETE /admin/brands/{id}", "Xóa thương hiệu"),
            # Admin Spec
            ("24", "Quản lý spec sản phẩm", "POST /admin/products/{id}/specs", "Thêm thông số kỹ thuật sản phẩm"),
            ("25", "Cập nhật spec sản phẩm", "PUT /admin/products/{id}/specs/{specId}", "Cập nhật thông số kỹ thuật"),
            ("26", "Xóa spec sản phẩm", "DELETE /admin/products/{id}/specs/{specId}", "Xóa thông số kỹ thuật"),
            # Excel Import
            ("27", "Import sản phẩm từ Excel", "POST /admin/products/excel/import", "Import sản phẩm hàng loạt từ file Excel"),
            ("28", "Download template Excel", "GET /admin/products/excel/template", "Tải file mẫu Excel để import"),
        ]
    },
    {
        "service": "cart-service",
        "color": "FFC000",  # vàng cam
        "features": [
            ("1", "Lấy giỏ hàng", "GET /cart", "Lấy danh sách sản phẩm trong giỏ hàng"),
            ("2", "Lấy giỏ hàng (chi tiết)", "GET /cart/items", "Lấy giỏ hàng kèm tổng tiền và số lượng"),
            ("3", "Xem tất cả giỏ hàng (Admin)", "GET /carts/admin", "Admin xem tất cả giỏ hàng đang tồn tại"),
            ("4", "Xem giỏ hàng theo ID (Admin)", "GET /carts/admin/{cartId}", "Admin xem giỏ hàng cụ thể"),
            ("5", "Thêm sản phẩm vào giỏ", "POST /cart?productId=&quantity=", "Thêm hoặc tăng số lượng sản phẩm trong giỏ"),
            ("6", "Cập nhật số lượng sản phẩm", "PUT /cart?productId=&quantity=", "Thay đổi số lượng sản phẩm trong giỏ"),
            ("7", "Xóa sản phẩm khỏi giỏ", "DELETE /cart?productId=", "Xóa một sản phẩm khỏi giỏ hàng"),
            ("8", "Xóa toàn bộ giỏ hàng", "DELETE /cart/clear", "Xóa tất cả sản phẩm trong giỏ"),
            ("9", "Xóa các sản phẩm đã chọn", "DELETE /cart/selected", "Xóa các sản phẩm được chọn để thanh toán"),
            ("10", "Xóa sản phẩm đã chọn (POST)", "POST /cart/selected/remove", "Alternative POST để xóa sản phẩm đã chọn"),
        ]
    },
    {
        "service": "order-service",
        "color": "FF0000",  # đỏ
        "features": [
            ("1", "Tạo đơn hàng từ giỏ hàng", "POST /order/{userId}", "Tạo đơn từ giỏ hàng hiện tại của user"),
            ("2", "Danh sách đơn hàng", "GET /orders", "Lấy tất cả đơn hàng (hỗ trợ filter)"),
            ("3", "Đơn hàng theo người dùng", "GET /users/{userId}/orders", "Lấy danh sách đơn hàng của một user"),
            ("4", "Tìm kiếm đơn hàng phân trang", "GET /orders/page", "Tìm kiếm & phân trang đơn hàng"),
            ("5", "Lấy chi tiết đơn hàng", "GET /orders/{id}", "Xem chi tiết một đơn hàng"),
            ("6", "Tra cứu đơn hàng (không đăng nhập)", "GET /orders/check?mvd=&phoneLast4=", "Tra cứu đơn hàng bằng mã MVD và 4 số cuối SĐT"),
            ("7", "Tạo đơn hàng thủ công (Admin)", "POST /orders/manual", "Admin tạo đơn hàng thay cho khách"),
            ("8", "Checkout COD - Mua ngay", "POST /orders/checkout/cod/buy-now", "Đặt hàng COD không qua giỏ hàng"),
            ("9", "Checkout COD - Từ giỏ hàng", "POST /orders/checkout/cod/cart", "Đặt hàng COD từ giỏ hàng"),
            ("10", "Cập nhật trạng thái đơn hàng", "PATCH /orders/{id}/status", "Admin cập nhật trạng thái xử lý đơn"),
            ("11", "Cập nhật trạng thái thanh toán", "PATCH/POST /orders/{id}/payment-status", "Cập nhật trạng thái thanh toán đơn hàng"),
            ("12", "Áp dụng voucher giảm giá", "- (internal)", "Áp dụng mã voucher khi tạo đơn hàng"),
            ("13", "Gửi email xác nhận đơn hàng", "- (internal)", "Gửi email thông báo đơn hàng thành công"),
        ]
    },
    {
        "service": "payment-service",
        "color": "7030A0",  # tím
        "features": [
            ("1", "Tạo bản ghi thanh toán", "POST /create", "Tạo payment record cho đơn hàng"),
            ("2", "Lấy thanh toán theo ID", "GET /{id}", "Lấy thông tin thanh toán theo ID"),
            ("3", "Lấy thanh toán theo đơn hàng", "GET /order/{orderId}", "Lấy danh sách thanh toán của đơn hàng"),
            ("4", "Đối soát thanh toán đã xử lý", "POST /order/{orderId}/reconcile-paid", "Đồng bộ lại trạng thái PAID với order-service"),
            ("5", "Hoàn tất thanh toán", "POST /{id}/complete", "Đánh dấu thanh toán thành công"),
            ("6", "Đánh dấu thanh toán thất bại", "POST /{id}/fail", "Đánh dấu thanh toán thất bại"),
        ]
    },
    {
        "service": "inventory-service",
        "color": "ED7D31",  # cam
        "features": [
            ("1", "Nhập kho (Inbound)", "POST /admin/stock/inbound", "Nhập hàng vào kho, tăng số lượng tồn kho"),
            ("2", "Xuất kho (Outbound)", "POST /admin/stock/outbound", "Xuất hàng, giảm số lượng tồn kho"),
            ("3", "Xem tồn kho theo sản phẩm", "GET /admin/stock/balance/{productId}", "Xem số lượng tồn kho theo sản phẩm"),
            ("4", "Xem tất cả tồn kho (phân trang)", "GET /admin/stock/balances", "Xem toàn bộ kho hàng, phân trang"),
            ("5", "Lịch sử xuất nhập theo SP", "GET /admin/stock/movements/{productId}", "Xem log nhập/xuất theo sản phẩm"),
            ("6", "Tất cả lịch sử xuất nhập", "GET /admin/stock/movements", "Xem toàn bộ log nhập/xuất kho"),
            ("7", "Tải template Excel nhập kho", "GET /admin/stock/excel/template", "Download file Excel mẫu nhập kho"),
            ("8", "Preview file Excel nhập kho", "POST /admin/stock/excel/preview", "Xem trước dữ liệu từ file Excel"),
            ("9", "Xác nhận nhập kho từ Excel", "POST /admin/stock/excel/confirm", "Xác nhận và lưu dữ liệu nhập kho từ Excel"),
        ]
    },
    {
        "service": "sale-service",
        "color": "C00000",  # đỏ đậm
        "features": [
            # Admin Sale Programs
            ("1", "Danh sách chương trình khuyến mãi (Admin)", "GET /admin/sales/programs", "Lấy tất cả chương trình sale"),
            ("2", "Chi tiết chương trình sale (Admin)", "GET /admin/sales/programs/{id}", "Xem chi tiết 1 chương trình"),
            ("3", "Tạo chương trình khuyến mãi", "POST /admin/sales/programs", "Tạo chương trình sale mới"),
            ("4", "Cập nhật chương trình sale", "PUT /admin/sales/programs/{id}", "Cập nhật thông tin chương trình"),
            ("5", "Xóa chương trình sale", "DELETE /admin/sales/programs/{id}", "Xóa chương trình khuyến mãi"),
            ("6", "Kiểm tra trùng thời gian sale", "GET /admin/sales/programs/check-overlap", "Kiểm tra sản phẩm có sale trùng thời điểm"),
            # Vouchers
            ("7", "Danh sách voucher (Admin)", "GET /admin/sales/vouchers", "Lấy tất cả voucher"),
            ("8", "Kiểm tra mã voucher trùng", "GET /admin/sales/vouchers/check-code", "Kiểm tra mã voucher đã tồn tại"),
            ("9", "Chi tiết voucher (Admin)", "GET /admin/sales/vouchers/{id}", "Xem chi tiết voucher"),
            ("10", "Tạo voucher mới", "POST /admin/sales/vouchers", "Tạo voucher giảm giá"),
            ("11", "Cập nhật voucher", "PUT /admin/sales/vouchers/{id}", "Cập nhật thông tin voucher"),
            ("12", "Xóa voucher", "DELETE /admin/sales/vouchers/{id}", "Xóa voucher"),
            # Banners
            ("13", "Danh sách banner (Admin)", "GET /admin/sales/banners", "Lấy tất cả banner quảng cáo"),
            ("14", "Chi tiết banner (Admin)", "GET /admin/sales/banners/{id}", "Xem chi tiết banner"),
            ("15", "Tạo banner quảng cáo", "POST /admin/sales/banners", "Tạo banner mới"),
            ("16", "Cập nhật banner", "PUT /admin/sales/banners/{id}", "Cập nhật thông tin banner"),
            ("17", "Xóa banner", "DELETE /admin/sales/banners/{id}", "Xóa banner"),
            ("18", "Upload hình banner", "POST /admin/sales/banners/upload-image", "Upload ảnh cho banner"),
            # Public
            ("19", "Sale đang chạy (Public)", "GET /sales/active", "Lấy các chương trình sale đang hoạt động"),
            ("20", "Banner đang active (Public)", "GET /sales/banners", "Lấy banner đang hiển thị"),
            ("21", "Xác thực voucher (Public)", "POST /sales/vouchers/validate", "Kiểm tra voucher hợp lệ không consume"),
            ("22", "Sử dụng voucher", "POST /sales/vouchers/consume", "Consume voucher sau khi đặt hàng thành công"),
        ]
    },
    {
        "service": "review-service",
        "color": "00B0F0",  # xanh nhạt
        "features": [
            ("1", "Viết đánh giá sản phẩm", "POST /reviews", "Tạo review mới cho sản phẩm đã mua"),
            ("2", "Chỉnh sửa đánh giá", "PUT /reviews/{reviewId}", "Cập nhật nội dung đánh giá"),
            ("3", "Lấy đánh giá theo sản phẩm", "GET /reviews/product/{productId}", "Xem tất cả đánh giá của sản phẩm"),
            ("4", "Đánh giá của tôi", "GET /reviews/user", "Xem đánh giá của người dùng hiện tại"),
            ("5", "Trả lời đánh giá", "POST /reviews/{reviewId}/reply", "Admin/người bán trả lời đánh giá"),
        ]
    },
    {
        "service": "notification-service",
        "color": "92D050",  # xanh lá nhạt
        "features": [
            ("1", "Gửi thông báo", "POST /send", "Gửi thông báo qua email/SMS/push"),
            ("2", "Lấy thông báo theo ID", "GET /{id}", "Xem chi tiết thông báo"),
            ("3", "Tra cứu thông báo theo email", "GET /lookup?email=", "Tìm tất cả thông báo của email"),
            ("4", "Đánh dấu đã đọc", "PATCH /{id}/read", "Đánh dấu thông báo là đã đọc"),
        ]
    },
    {
        "service": "ai-chatbot-service",
        "color": "00B0F0",  # xanh dương sáng
        "features": [
            ("1", "Chat với AI", "POST /chat", "Gửi tin nhắn và nhận phản hồi từ AI tư vấn sản phẩm"),
        ]
    },
    {
        "service": "activity-log-service",
        "color": "595959",  # xám
        "features": [
            ("1", "Ghi nhật ký hoạt động", "POST /log", "Ghi lại hoạt động hệ thống vào database"),
            ("2", "Xem hoạt động gần đây", "GET /recent?limit=", "Xem N hoạt động gần nhất"),
            ("3", "Lấy log theo ID", "GET /{id}", "Xem chi tiết log theo ID"),
            ("4", "Xóa nhiều log", "DELETE /batch", "Xóa nhiều bản ghi log cùng lúc"),
        ]
    },
    {
        "service": "telegram-service",
        "color": "0563C1",  # xanh Telegram
        "features": [
            ("1", "Liên kết tài khoản Telegram", "POST /telegram/auth/link", "Liên kết tài khoản với Telegram bot"),
            ("2", "Hủy liên kết Telegram", "POST /telegram/auth/unlink", "Hủy liên kết tài khoản Telegram"),
            ("3", "Telegram bot admin commands", "- (bot handler)", "Nhận lệnh quản trị qua Telegram bot"),
            ("4", "Chỉnh sửa sản phẩm qua Telegram", "- (bot handler)", "Admin chỉnh sửa giá/thông tin SP qua Telegram"),
        ]
    },
    {
        "service": "product-recommendation-service",
        "color": "FF5F00",  # cam đậm
        "features": [
            ("1", "Lấy sản phẩm gợi ý", "GET /recommendations/{userId}", "Lấy sản phẩm gợi ý cá nhân hóa theo user"),
            ("2", "Sản phẩm tương tự", "GET /recommendations/similar/{productId}", "Lấy sản phẩm tương tự"),
        ]
    },
]


def hex_to_argb(hex_color):
    return "FF" + hex_color.upper()


def create_test_plan_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # =====================================================================
    # SHEET TỔNG QUAN
    # =====================================================================
    summary_ws = wb.create_sheet("📊 Tổng Quan", 0)

    # Header gradient
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")

    summary_ws.merge_cells("A1:H1")
    summary_ws["A1"] = "📋 BẢNG TIẾN ĐỘ TEST CHỨC NĂNG - E-COMMERCE MICROSERVICES"
    summary_ws["A1"].fill = header_fill
    summary_ws["A1"].font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_ws.row_dimensions[1].height = 45

    # Sub header
    summary_ws.merge_cells("A2:H2")
    summary_ws["A2"] = f"Dự án: E-Commerce Microservices  |  Ngày tạo: 26/06/2026  |  Phiên bản: 1.0"
    summary_ws["A2"].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    summary_ws["A2"].font = Font(color="FFFFFF", size=11, italic=True, name="Calibri")
    summary_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[2].height = 25

    # Column headers
    col_headers = [
        ("A", "STT", 6),
        ("B", "TÊN SERVICE", 30),
        ("C", "TỔNG CHỨC NĂNG", 18),
        ("D", "ĐÃ TEST", 12),
        ("E", "CHƯA TEST", 12),
        ("F", "ĐẠT (PASS)", 12),
        ("G", "CHƯA ĐẠT (FAIL)", 15),
        ("H", "% HOÀN THÀNH", 15),
    ]

    col_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    col_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")

    for col, header, width in col_headers:
        cell = summary_ws[f"{col}3"]
        cell.value = header
        cell.fill = col_fill
        cell.font = col_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        summary_ws.column_dimensions[col].width = width

    summary_ws.row_dimensions[3].height = 35

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    for i, svc in enumerate(SERVICES_DATA, start=1):
        row = i + 3
        total = len(svc["features"])
        summary_ws.row_dimensions[row].height = 22

        cells_data = [
            ("A", i, "center"),
            ("B", svc["service"], "left"),
            ("C", total, "center"),
            ("D", 0, "center"),  # Đã test
            ("E", total, "center"),  # Chưa test
            ("F", 0, "center"),  # Pass
            ("G", 0, "center"),  # Fail
            ("H", "0%", "center"),  # %
        ]

        row_fill = PatternFill(
            start_color=svc["color"],
            end_color=svc["color"],
            fill_type="solid"
        ) if i % 2 == 1 else PatternFill(
            start_color="F2F2F2",
            end_color="F2F2F2",
            fill_type="solid"
        )

        # Alternate row colors based on service color
        alt_fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")

        for col, value, align in cells_data:
            cell = summary_ws[f"{col}{row}"]
            cell.value = value
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = thin_border
            if col == "B":
                cell.fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
                cell.font = Font(bold=True, color=svc["color"], name="Calibri", size=11)
            else:
                cell.fill = alt_fill if i % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.font = Font(name="Calibri", size=10)

    # Total row
    total_row = len(SERVICES_DATA) + 4
    summary_ws.merge_cells(f"A{total_row}:B{total_row}")
    summary_ws[f"A{total_row}"] = "TỔNG CỘNG"
    summary_ws[f"A{total_row}"].font = Font(bold=True, name="Calibri", size=11)
    summary_ws[f"A{total_row}"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    summary_ws[f"A{total_row}"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    summary_ws[f"A{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
    summary_ws[f"A{total_row}"].border = thin_border

    total_features = sum(len(svc["features"]) for svc in SERVICES_DATA)
    total_cell = summary_ws[f"C{total_row}"]
    total_cell.value = total_features
    total_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    total_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.border = thin_border
    summary_ws.row_dimensions[total_row].height = 28

    for col in ["D", "E", "F", "G", "H"]:
        c = summary_ws[f"{col}{total_row}"]
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        if col == "E":
            c.value = total_features
        else:
            c.value = 0 if col != "H" else "0%"

    # Legend
    legend_row = total_row + 2
    summary_ws.merge_cells(f"A{legend_row}:H{legend_row}")
    summary_ws[f"A{legend_row}"] = "BẢNG MÀU TRẠNG THÁI"
    summary_ws[f"A{legend_row}"].font = Font(bold=True, size=11, name="Calibri")
    summary_ws[f"A{legend_row}"].alignment = Alignment(horizontal="left")

    legend_data = [
        ("PASS", "92D050", "Chức năng hoạt động đúng"),
        ("FAIL", "FF0000", "Chức năng lỗi, cần sửa"),
        ("CHƯA TEST", "BFBFBF", "Chưa thực hiện test"),
        ("BỎ QUA", "FFC000", "Tạm thời bỏ qua"),
    ]

    for idx, (status, color, desc) in enumerate(legend_data):
        r = legend_row + 1 + idx
        c1 = summary_ws[f"A{r}"]
        c1.value = status
        c1.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c1.font = Font(bold=True, color="FFFFFF" if color in ["FF0000", "BFBFBF", "1F3864"] else "000000", name="Calibri")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = thin_border
        summary_ws.row_dimensions[r].height = 22

        summary_ws.merge_cells(f"B{r}:H{r}")
        c2 = summary_ws[f"B{r}"]
        c2.value = desc
        c2.font = Font(name="Calibri", size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = thin_border

    # =====================================================================
    # TẠO SHEET CHO TỪNG SERVICE
    # =====================================================================
    STATUS_OPTS = ["CHƯA TEST", "PASS", "FAIL", "BỎ QUA"]

    for svc in SERVICES_DATA:
        sname = svc["service"]
        color = svc["color"]
        # Sheet name: max 31 chars
        sheet_name = sname[:28] if len(sname) > 28 else sname
        ws = wb.create_sheet(f"🔧 {sheet_name}")

        svc_color = hex_to_argb(color)

        # ---- HEADER ----
        ws.merge_cells("A1:K1")
        ws["A1"] = f"📋 {sname.upper()} — BẢNG TIẾN ĐỘ TEST CHỨC NĂNG"
        ws["A1"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:K2")
        ws["A2"] = f"Service: {sname}  |  Tổng số chức năng: {len(svc['features'])}  |  Ngày cập nhật: 26/06/2026"
        ws["A2"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws["A2"].font = Font(color="FFFFFF", size=11, italic=True, name="Calibri")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # ---- COLUMN HEADERS ----
        headers = [
            ("A", "STT", 6),
            ("B", "TÊN CHỨC NĂNG", 38),
            ("C", "API ENDPOINT", 42),
            ("D", "MÔ TẢ", 45),
            ("E", "LOẠI TEST", 15),
            ("F", "TRẠNG THÁI", 14),
            ("G", "KẾT QUẢ THỰC TẾ", 30),
            ("H", "LỖI / GHI CHÚ", 35),
            ("I", "NGƯỜI TEST", 18),
            ("J", "NGÀY TEST", 14),
            ("K", "ƯU TIÊN", 10),
        ]

        col_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        col_header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")

        for col, header, width in headers:
            cell = ws[f"{col}3"]
            cell.value = header
            cell.fill = col_header_fill
            cell.font = col_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[col].width = width

        ws.row_dimensions[3].height = 35

        # ---- DATA ROWS ----
        for ridx, (stt, fname, endpoint, desc) in enumerate(svc["features"], start=1):
            row = ridx + 3

            # Alternate row fill
            if ridx % 2 == 0:
                base_fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
            else:
                base_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            ws.row_dimensions[row].height = 22

            data = [
                ("A", stt),
                ("B", fname),
                ("C", endpoint),
                ("D", desc),
                ("E", "API/Integration"),
                ("F", "CHƯA TEST"),
                ("G", ""),
                ("H", ""),
                ("I", ""),
                ("J", ""),
                ("K", "Cao" if ridx <= 3 else "Trung bình"),
            ]

            for col, value in data:
                cell = ws[f"{col}{row}"]
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center" if col in ["A", "E", "F", "I", "J", "K"] else "left",
                    vertical="center",
                    wrap_text=True
                )

                if col == "A":
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                elif col == "F":
                    cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
                    cell.font = Font(bold=True, size=9, name="Calibri")
                else:
                    cell.fill = base_fill
                    if col == "B":
                        cell.font = Font(bold=True, size=10, name="Calibri")
                    else:
                        cell.font = Font(size=10, name="Calibri")

        # Freeze header rows
        ws.freeze_panes = "B4"

        # Data validation dropdown for Status column
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"CHƯA TEST,PASS,FAIL,BỎ QUA"',
            allow_blank=False,
            showDropDown=False
        )
        dv.error = "Chỉ chọn: CHƯA TEST, PASS, FAIL, hoặc BỎ QUA"
        dv.errorTitle = "Giá trị không hợp lệ"
        dv.prompt = "Chọn trạng thái test"
        dv.promptTitle = "Trạng thái"
        ws.add_data_validation(dv)
        dv.sqref = f"F4:F{3 + len(svc['features'])}"

        # Data validation for Priority
        dv2 = DataValidation(
            type="list",
            formula1='"Cao,Trung bình,Thấp"',
            allow_blank=False,
            showDropDown=False
        )
        ws.add_data_validation(dv2)
        dv2.sqref = f"K4:K{3 + len(svc['features'])}"

        # Data validation for Test type
        dv3 = DataValidation(
            type="list",
            formula1='"API/Integration,Unit,Manual,E2E"',
            allow_blank=False,
            showDropDown=False
        )
        ws.add_data_validation(dv3)
        dv3.sqref = f"E4:E{3 + len(svc['features'])}"

    # =====================================================================
    # SHEET HƯỚNG DẪN
    # =====================================================================
    guide_ws = wb.create_sheet("📖 Hướng Dẫn")
    guide_ws.column_dimensions["A"].width = 20
    guide_ws.column_dimensions["B"].width = 60

    guide_ws.merge_cells("A1:B1")
    guide_ws["A1"] = "📖 HƯỚNG DẪN SỬ DỤNG BẢNG TIẾN ĐỘ TEST"
    guide_ws["A1"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    guide_ws["A1"].font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    guide_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    guide_ws.row_dimensions[1].height = 40

    guide_data = [
        ("", ""),
        ("🎯 MỤC ĐÍCH", "File Excel này dùng để theo dõi tiến độ test từng chức năng của mỗi microservice"),
        ("", ""),
        ("📝 CÁC CỘT", ""),
        ("STT", "Số thứ tự chức năng trong service"),
        ("TÊN CHỨC NĂNG", "Tên mô tả ngắn gọn chức năng cần test"),
        ("API ENDPOINT", "Đường dẫn API (Method + URL)"),
        ("MÔ TẢ", "Mô tả chi tiết hơn về chức năng"),
        ("LOẠI TEST", "Chọn: API/Integration, Unit, Manual, E2E"),
        ("TRẠNG THÁI", "CHƯA TEST / PASS / FAIL / BỎ QUA"),
        ("KẾT QUẢ THỰC TẾ", "Ghi rõ kết quả khi test (response, data trả về...)"),
        ("LỖI / GHI CHÚ", "Mô tả lỗi nếu FAIL, hoặc ghi chú đặc biệt"),
        ("NGƯỜI TEST", "Tên người thực hiện test"),
        ("NGÀY TEST", "Ngày thực hiện test (dd/mm/yyyy)"),
        ("ƯU TIÊN", "Cao / Trung bình / Thấp"),
        ("", ""),
        ("🚦 TRẠNG THÁI", ""),
        ("CHƯA TEST", "Chức năng chưa được test (màu xám)"),
        ("PASS", "Test thành công, chức năng hoạt động đúng (màu xanh lá)"),
        ("FAIL", "Test thất bại, cần fix lỗi (màu đỏ)"),
        ("BỎ QUA", "Tạm thời bỏ qua, không test trong đợt này (màu vàng)"),
        ("", ""),
        ("📊 TỔNG QUAN", "Sheet 'Tổng Quan' hiển thị thống kê tổng hợp của tất cả service"),
        ("", ""),
        ("💡 TIPS", "Dùng Ctrl+F để tìm kiếm nhanh chức năng cần test"),
        ("", "Cập nhật cột NGÀY TEST ngay sau khi test xong"),
        ("", "Ghi rõ lỗi ở cột 'LỖI / GHI CHÚ' để developer dễ fix"),
    ]

    for ridx, (key, value) in enumerate(guide_data, start=2):
        r = ridx
        guide_ws.row_dimensions[r].height = 22
        k_cell = guide_ws[f"A{r}"]
        v_cell = guide_ws[f"B{r}"]
        k_cell.value = key
        v_cell.value = value
        k_cell.font = Font(bold=True, name="Calibri", size=10)
        v_cell.font = Font(name="Calibri", size=10)
        v_cell.alignment = Alignment(wrap_text=True, vertical="center")

        if key and not key.startswith("📝") and not key.startswith("🚦") and not key.startswith("📊") and not key.startswith("🎯") and not key.startswith("💡"):
            k_cell.fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
            k_cell.border = thin_border
            v_cell.border = thin_border

    # Save the workbook
    output_path = "TEST_TIEN_DO_MICROSERVICES.xlsx"
    wb.save(output_path)
    total = sum(len(s["features"]) for s in SERVICES_DATA)
    print("[OK] Da tao file Excel: " + output_path)
    print(f"[OK] Tong so chuc nang: {total} tren {len(SERVICES_DATA)} services")


if __name__ == "__main__":
    create_test_plan_excel()
